"""
Excel Tool

Manipulate Excel (.xlsx) files locally.

Requires:
    pip install strands-pack[excel]

Supported actions
-----------------
- create_workbook
    Parameters: output_path (required), sheet_name (default "Sheet1")
- read_workbook
    Parameters: input_path (required), sheet_name (optional), data_only (default True)
- read_range
    Parameters: input_path (required), range (required, e.g., "A1:C10"), sheet_name (optional)
- write_range
    Parameters: input_path (required), range (required), values (required - 2D list), sheet_name (optional)
- add_sheet
    Parameters: input_path (required), sheet_name (required), position (optional)
- delete_sheet
    Parameters: input_path (required), sheet_name (required)
- list_sheets
    Parameters: input_path (required)
- get_info
    Parameters: input_path (required)
- apply_formula
    Parameters: input_path (required), cell (required, e.g., "A1"), formula (required), sheet_name (optional)
- save_as
    Parameters: input_path (required), output_path (required)

Notes:
  - All operations preserve existing data unless explicitly modified
  - data_only=True returns calculated values, False returns formulas
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

from strands import tool

# Lazy import for openpyxl
_openpyxl = None


def _get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        try:
            import openpyxl
            _openpyxl = openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install strands-pack[excel]") from None
    return _openpyxl


def _ok(**data: Any) -> Dict[str, Any]:
    out: Dict[str, Any] = {"success": True}
    out.update(data)
    return out


def _err(message: str, *, error_type: Optional[str] = None, **data: Any) -> Dict[str, Any]:
    out: Dict[str, Any] = {"success": False, "error": message}
    if error_type:
        out["error_type"] = error_type
    out.update(data)
    return out


def _cell_value(cell) -> Any:
    """Convert cell value to JSON-serializable format."""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, (int, float, str, bool)):
        return val
    # Handle datetime
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def _create_workbook(output_path: str, sheet_name: str = "Sheet1", **kwargs) -> Dict[str, Any]:
    """Create a new Excel workbook."""
    if not output_path:
        return _err("output_path is required")

    openpyxl = _get_openpyxl()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    path = Path(output_path).expanduser()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()

    return _ok(
        action="create_workbook",
        output_path=str(path),
        sheet_name=sheet_name,
    )


def _read_workbook(input_path: str, sheet_name: Optional[str] = None,
                   data_only: bool = True, **kwargs) -> Dict[str, Any]:
    """Read entire workbook or specific sheet."""
    if not input_path:
        return _err("input_path is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path), data_only=data_only)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return _err(f"Sheet not found: {sheet_name}", available_sheets=wb.sheetnames)
        sheets_to_read = [sheet_name]
    else:
        sheets_to_read = wb.sheetnames

    result: Dict[str, List[List[Any]]] = {}
    for sn in sheets_to_read:
        ws = wb[sn]
        data = []
        for row in ws.iter_rows():
            row_data = [_cell_value(cell) for cell in row]
            data.append(row_data)
        result[sn] = data

    wb.close()

    return _ok(
        action="read_workbook",
        input_path=str(path),
        sheets=result,
        sheet_count=len(result),
    )


def _read_range(input_path: str, range: str, sheet_name: Optional[str] = None,
                data_only: bool = True, **kwargs) -> Dict[str, Any]:
    """Read a specific range from a sheet."""
    if not input_path:
        return _err("input_path is required")
    if not range:
        return _err("range is required (e.g., 'A1:C10')")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path), data_only=data_only)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return _err(f"Sheet not found: {sheet_name}", available_sheets=wb.sheetnames)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    try:
        cells = ws[range]
    except Exception as e:
        wb.close()
        return _err(f"Invalid range: {range}", error_type=type(e).__name__)

    # Handle single cell vs range
    if not hasattr(cells, "__iter__") or isinstance(cells, str):
        data = [[_cell_value(cells)]]
    elif hasattr(cells, "value"):
        # Single cell
        data = [[_cell_value(cells)]]
    else:
        # Range of cells (tuple of tuples)
        data = []
        for row in cells:
            if hasattr(row, "__iter__") and not isinstance(row, str):
                row_data = [_cell_value(cell) for cell in row]
            else:
                row_data = [_cell_value(row)]
            data.append(row_data)

    wb.close()

    return _ok(
        action="read_range",
        input_path=str(path),
        sheet_name=sheet_name,
        range=range,
        data=data,
        rows=len(data),
        cols=len(data[0]) if data else 0,
    )


def _write_range(input_path: str, range: str, values: List[List[Any]],
                 sheet_name: Optional[str] = None, **kwargs) -> Dict[str, Any]:
    """Write values to a specific range."""
    if not input_path:
        return _err("input_path is required")
    if not range:
        return _err("range is required (e.g., 'A1')")
    if not values:
        return _err("values is required (2D list)")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    wb = openpyxl.load_workbook(str(path))

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return _err(f"Sheet not found: {sheet_name}", available_sheets=wb.sheetnames)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    # Parse starting cell from range (e.g., "A1" or "A1:C10" -> start at A1)
    start_cell = range.split(":")[0]
    try:
        col_letter, row_num = coordinate_from_string(start_cell)
        start_col = column_index_from_string(col_letter)
        start_row = row_num
    except Exception as e:
        wb.close()
        return _err(f"Invalid range: {range}", error_type=type(e).__name__)

    # Write values
    cells_written = 0
    for row_idx, row_values in enumerate(values):
        for col_idx, value in enumerate(row_values):
            ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
            cells_written += 1

    wb.save(str(path))
    wb.close()

    return _ok(
        action="write_range",
        input_path=str(path),
        sheet_name=sheet_name,
        range=range,
        rows_written=len(values),
        cols_written=len(values[0]) if values else 0,
        cells_written=cells_written,
    )


def _add_sheet(input_path: str, sheet_name: str, position: Optional[int] = None,
               **kwargs) -> Dict[str, Any]:
    """Add a new sheet to the workbook."""
    if not input_path:
        return _err("input_path is required")
    if not sheet_name:
        return _err("sheet_name is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path))

    if sheet_name in wb.sheetnames:
        wb.close()
        return _err(f"Sheet already exists: {sheet_name}")

    if position is not None:
        wb.create_sheet(sheet_name, position)
    else:
        wb.create_sheet(sheet_name)

    wb.save(str(path))
    wb.close()

    return _ok(
        action="add_sheet",
        input_path=str(path),
        sheet_name=sheet_name,
        position=position,
        all_sheets=wb.sheetnames,
    )


def _delete_sheet(input_path: str, sheet_name: str, **kwargs) -> Dict[str, Any]:
    """Delete a sheet from the workbook."""
    if not input_path:
        return _err("input_path is required")
    if not sheet_name:
        return _err("sheet_name is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path))

    if sheet_name not in wb.sheetnames:
        wb.close()
        return _err(f"Sheet not found: {sheet_name}", available_sheets=wb.sheetnames)

    if len(wb.sheetnames) == 1:
        wb.close()
        return _err("Cannot delete the only sheet in the workbook")

    del wb[sheet_name]

    wb.save(str(path))
    remaining_sheets = wb.sheetnames
    wb.close()

    return _ok(
        action="delete_sheet",
        input_path=str(path),
        deleted_sheet=sheet_name,
        remaining_sheets=remaining_sheets,
    )


def _list_sheets(input_path: str, **kwargs) -> Dict[str, Any]:
    """List all sheets in the workbook."""
    if not input_path:
        return _err("input_path is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path), read_only=True)
    sheets = wb.sheetnames
    wb.close()

    return _ok(
        action="list_sheets",
        input_path=str(path),
        sheets=sheets,
        count=len(sheets),
    )


def _get_info(input_path: str, **kwargs) -> Dict[str, Any]:
    """Get information about the workbook."""
    if not input_path:
        return _err("input_path is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path), read_only=True)

    sheets_info = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets_info.append({
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })

    file_size = path.stat().st_size

    wb.close()

    return _ok(
        action="get_info",
        input_path=str(path),
        file_size_bytes=file_size,
        sheet_count=len(sheets_info),
        sheets=sheets_info,
    )


def _apply_formula(input_path: str, cell: str, formula: str,
                   sheet_name: Optional[str] = None, **kwargs) -> Dict[str, Any]:
    """Apply a formula to a cell."""
    if not input_path:
        return _err("input_path is required")
    if not cell:
        return _err("cell is required (e.g., 'A1')")
    if not formula:
        return _err("formula is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path))

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return _err(f"Sheet not found: {sheet_name}", available_sheets=wb.sheetnames)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    # Ensure formula starts with =
    if not formula.startswith("="):
        formula = "=" + formula

    ws[cell] = formula

    wb.save(str(path))
    wb.close()

    return _ok(
        action="apply_formula",
        input_path=str(path),
        sheet_name=sheet_name,
        cell=cell,
        formula=formula,
    )


def _save_as(input_path: str, output_path: str, **kwargs) -> Dict[str, Any]:
    """Save workbook to a new path."""
    if not input_path:
        return _err("input_path is required")
    if not output_path:
        return _err("output_path is required")

    path = Path(input_path).expanduser()
    if not path.exists():
        return _err(f"File not found: {input_path}", error_type="FileNotFoundError")

    openpyxl = _get_openpyxl()

    wb = openpyxl.load_workbook(str(path))

    out_path = Path(output_path).expanduser()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(str(out_path))
    wb.close()

    return _ok(
        action="save_as",
        input_path=str(path),
        output_path=str(out_path),
    )


_ACTIONS = {
    "create_workbook": _create_workbook,
    "read_workbook": _read_workbook,
    "read_range": _read_range,
    "write_range": _write_range,
    "add_sheet": _add_sheet,
    "delete_sheet": _delete_sheet,
    "list_sheets": _list_sheets,
    "get_info": _get_info,
    "apply_formula": _apply_formula,
    "save_as": _save_as,
}


@tool
def excel(
    action: str,
    input_path: Optional[str] = None,
    output_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    data_only: bool = True,
    range: Optional[str] = None,
    values: Optional[List[List[Any]]] = None,
    position: Optional[int] = None,
    cell: Optional[str] = None,
    formula: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Manipulate Excel (.xlsx) files locally.

    Actions:
    - create_workbook: Create a new Excel workbook
    - read_workbook: Read entire workbook or specific sheet
    - read_range: Read a specific cell range
    - write_range: Write values to a cell range
    - add_sheet: Add a new sheet
    - delete_sheet: Delete a sheet
    - list_sheets: List all sheets
    - get_info: Get workbook information
    - apply_formula: Apply a formula to a cell
    - save_as: Save workbook to a new path

    Args:
        action: The action to perform (required)
        input_path: Path to the input Excel file (required for most actions)
        output_path: Path for output file (required for create_workbook, save_as)
        sheet_name: Name of the sheet (optional for some actions, required for add_sheet/delete_sheet)
        data_only: If True, return calculated values; if False, return formulas (default True)
        range: Cell range like "A1:C10" (required for read_range, write_range)
        values: 2D list of values to write (required for write_range)
        position: Position index for new sheet (optional for add_sheet)
        cell: Cell reference like "A1" (required for apply_formula)
        formula: Formula to apply (required for apply_formula)

    Returns:
        dict with success status and action-specific data
    """
    action = (action or "").strip().lower()

    if action not in _ACTIONS:
        return _err(
            f"Unknown action: {action}",
            error_type="InvalidAction",
            available_actions=list(_ACTIONS.keys()),
        )

    # Build kwargs dict from explicit parameters
    kwargs: Dict[str, Any] = {}
    if input_path is not None:
        kwargs["input_path"] = input_path
    if output_path is not None:
        kwargs["output_path"] = output_path
    if sheet_name is not None:
        kwargs["sheet_name"] = sheet_name
    if data_only is not True:  # Only pass if not the default
        kwargs["data_only"] = data_only
    else:
        kwargs["data_only"] = data_only
    if range is not None:
        kwargs["range"] = range
    if values is not None:
        kwargs["values"] = values
    if position is not None:
        kwargs["position"] = position
    if cell is not None:
        kwargs["cell"] = cell
    if formula is not None:
        kwargs["formula"] = formula

    try:
        return _ACTIONS[action](**kwargs)
    except ImportError as e:
        return _err(str(e), error_type="ImportError")
    except Exception as e:
        return _err(str(e), error_type=type(e).__name__, action=action)
