"""Tests for Excel tool."""

import os
import tempfile

import pytest


@pytest.fixture
def output_dir():
    """Create a temp directory for output files."""
    with tempfile.TemporaryDirectory() as d:
        yield d


@pytest.fixture
def sample_workbook(output_dir):
    """Create a sample Excel workbook for testing."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some data
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Test1"
    ws["B2"] = 100
    ws["A3"] = "Test2"
    ws["B3"] = 200

    path = os.path.join(output_dir, "sample.xlsx")
    wb.save(path)
    wb.close()
    return path


def test_excel_create_workbook(output_dir):
    """Test creating a new workbook."""
    from strands_pack import excel

    output_path = os.path.join(output_dir, "new_workbook.xlsx")
    result = excel(action="create_workbook", output_path=output_path)

    assert result["success"] is True
    assert result["action"] == "create_workbook"
    assert os.path.exists(output_path)
    assert result["sheet_name"] == "Sheet1"


def test_excel_create_workbook_custom_sheet(output_dir):
    """Test creating a workbook with custom sheet name."""
    from strands_pack import excel

    output_path = os.path.join(output_dir, "custom_sheet.xlsx")
    result = excel(
        action="create_workbook",
        output_path=output_path,
        sheet_name="MyData",
    )

    assert result["success"] is True
    assert result["sheet_name"] == "MyData"


def test_excel_read_workbook(sample_workbook):
    """Test reading a workbook."""
    from strands_pack import excel

    result = excel(action="read_workbook", input_path=sample_workbook)

    assert result["success"] is True
    assert result["action"] == "read_workbook"
    assert "Sheet1" in result["sheets"]
    assert result["sheet_count"] == 1


def test_excel_read_range(sample_workbook):
    """Test reading a specific range."""
    from strands_pack import excel

    result = excel(
        action="read_range",
        input_path=sample_workbook,
        range="A1:B3",
    )

    assert result["success"] is True
    assert result["action"] == "read_range"
    assert result["rows"] == 3
    assert result["cols"] == 2
    assert result["data"][0][0] == "Name"
    assert result["data"][1][1] == 100


def test_excel_write_range(sample_workbook):
    """Test writing to a range."""
    from strands_pack import excel

    result = excel(
        action="write_range",
        input_path=sample_workbook,
        range="C1",
        values=[["Status"], ["Active"], ["Inactive"]],
    )

    assert result["success"] is True
    assert result["action"] == "write_range"
    assert result["rows_written"] == 3
    assert result["cols_written"] == 1

    # Verify the data was written
    read_result = excel(
        action="read_range",
        input_path=sample_workbook,
        range="C1:C3",
    )
    assert read_result["data"][0][0] == "Status"
    assert read_result["data"][1][0] == "Active"


def test_excel_add_sheet(sample_workbook):
    """Test adding a new sheet."""
    from strands_pack import excel

    result = excel(
        action="add_sheet",
        input_path=sample_workbook,
        sheet_name="NewSheet",
    )

    assert result["success"] is True
    assert result["action"] == "add_sheet"
    assert result["sheet_name"] == "NewSheet"


def test_excel_add_sheet_duplicate(sample_workbook):
    """Test error when adding duplicate sheet."""
    from strands_pack import excel

    result = excel(
        action="add_sheet",
        input_path=sample_workbook,
        sheet_name="Sheet1",
    )

    assert result["success"] is False
    assert "already exists" in result["error"]


def test_excel_delete_sheet(sample_workbook):
    """Test deleting a sheet."""
    from strands_pack import excel

    # First add a second sheet
    excel(action="add_sheet", input_path=sample_workbook, sheet_name="ToDelete")

    # Then delete it
    result = excel(
        action="delete_sheet",
        input_path=sample_workbook,
        sheet_name="ToDelete",
    )

    assert result["success"] is True
    assert result["action"] == "delete_sheet"
    assert "ToDelete" not in result["remaining_sheets"]


def test_excel_delete_only_sheet(sample_workbook):
    """Test error when deleting the only sheet."""
    from strands_pack import excel

    result = excel(
        action="delete_sheet",
        input_path=sample_workbook,
        sheet_name="Sheet1",
    )

    assert result["success"] is False
    assert "only sheet" in result["error"].lower()


def test_excel_list_sheets(sample_workbook):
    """Test listing sheets."""
    from strands_pack import excel

    result = excel(action="list_sheets", input_path=sample_workbook)

    assert result["success"] is True
    assert result["action"] == "list_sheets"
    assert "Sheet1" in result["sheets"]
    assert result["count"] == 1


def test_excel_get_info(sample_workbook):
    """Test getting workbook info."""
    from strands_pack import excel

    result = excel(action="get_info", input_path=sample_workbook)

    assert result["success"] is True
    assert result["action"] == "get_info"
    assert result["sheet_count"] == 1
    assert "file_size_bytes" in result


def test_excel_apply_formula(sample_workbook):
    """Test applying a formula."""
    from strands_pack import excel

    result = excel(
        action="apply_formula",
        input_path=sample_workbook,
        cell="C1",
        formula="=SUM(B2:B3)",
    )

    assert result["success"] is True
    assert result["action"] == "apply_formula"
    assert result["formula"] == "=SUM(B2:B3)"


def test_excel_save_as(sample_workbook, output_dir):
    """Test saving as a new file."""
    from strands_pack import excel

    new_path = os.path.join(output_dir, "copy.xlsx")
    result = excel(
        action="save_as",
        input_path=sample_workbook,
        output_path=new_path,
    )

    assert result["success"] is True
    assert result["action"] == "save_as"
    assert os.path.exists(new_path)


def test_excel_file_not_found():
    """Test error when file doesn't exist."""
    from strands_pack import excel

    result = excel(action="read_workbook", input_path="/nonexistent/file.xlsx")

    assert result["success"] is False
    assert "not found" in result["error"].lower()


def test_excel_missing_input_path():
    """Test error when input path is missing."""
    from strands_pack import excel

    result = excel(action="read_workbook")

    assert result["success"] is False
    assert "input_path" in result["error"]


def test_excel_unknown_action():
    """Test error for unknown action."""
    from strands_pack import excel

    result = excel(action="unknown_action")

    assert result["success"] is False
    assert "Unknown action" in result["error"]
    assert "available_actions" in result
